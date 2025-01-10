import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv
from convert_to_pdf import convert_all_excel_to_pdf
from sent_email import send_email_with_attachments
from delete_folder import delete_folder
from merger_pdf import merge_pdfs_in_folder


# Đường dẫn root
root_path = "D:/StockAutoReport"

symbols_file = 'symbols.txt'

end_date = datetime.now()
start_date = end_date - timedelta(weeks=4)

# Format ngày dạng dd/mm/yyyy
end_date_str = end_date.strftime("%d/%m/%Y")
start_date_str = start_date.strftime("%d/%m/%Y")

# Tạo chuỗi start_str và end_str để đặt tên folder
start_str = start_date.strftime("%d%m%Y")
end_str = end_date.strftime("%d%m%Y")

# Tạo folder output
xlsx_folder = os.path.join(root_path, f"xlsx_{start_str}_{end_str}")
pdf_folder = os.path.join(root_path, f"pdf_{start_str}_{end_str}")

os.makedirs(xlsx_folder, exist_ok=True)
os.makedirs(pdf_folder, exist_ok=True)

page_index = 1
page_size = 20

base_url = "https://cafef.vn/du-lieu/Ajax/PageNew/DataHistory/PriceHistory.ashx"

with open(symbols_file, 'r', encoding='utf-8') as f:
    symbols = [line.strip() for line in f if line.strip()]

for symbol in symbols:
    params = {
        "Symbol": symbol,
        "StartDate": start_date_str,
        "EndDate": end_date_str,
        "PageIndex": page_index,
        "PageSize": page_size
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        if isinstance(data, dict) and "Data" in data and "Data" in data["Data"]:
            records = data["Data"]["Data"]
        else:
            print(f"Dữ liệu trả về cho {symbol} không đúng định dạng.")
            continue
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = symbol
        
        # Cài đặt trang in
        ws.page_setup.orientation = 'landscape'
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        # Header
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Mã cổ phiếu: {symbol} - {start_date_str} - {end_date_str}"
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("A2:A3")
        ws["A2"].value = "Ngày"
        ws["A2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("B2:C2")
        ws["B2"].value = "Giá (nghìn VNĐ)"
        ws["B2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws["B3"].value = "Đóng cửa"
        ws["B3"].alignment = Alignment(horizontal='center', vertical='center')
        ws["C3"].value = "Điều chỉnh"
        ws["C3"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("D2:D3")
        ws["D2"].value = "Thay đổi"
        ws["D2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("E2:F2")
        ws["E2"].value = "GD khớp lệnh"
        ws["E2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws["E3"].value = "Khối lượng"
        ws["E3"].alignment = Alignment(horizontal='center', vertical='center')
        ws["F3"].value = "Giá trị (tỷ VNĐ)"
        ws["F3"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("G2:H2")
        ws["G2"].value = "GD thỏa thuận"
        ws["G2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws["G3"].value = "Khối lượng"
        ws["G3"].alignment = Alignment(horizontal='center', vertical='center')
        ws["H3"].value = "Giá trị (tỷ VNĐ)"
        ws["H3"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells("I2:K2")
        ws["I2"].value = "Giá (nghìn VNĐ)"
        ws["I2"].alignment = Alignment(horizontal='center', vertical='center')
        
        ws["I3"].value = "Mở cửa"
        ws["I3"].alignment = Alignment(horizontal='center', vertical='center')
        ws["J3"].value = "Cao nhất"
        ws["J3"].alignment = Alignment(horizontal='center', vertical='center')
        ws["K3"].value = "Thấp nhất"
        ws["K3"].alignment = Alignment(horizontal='center', vertical='center')
        # Dữ liệu từ hàng 4
        row_start = 4
        for i, record in enumerate(records, start=row_start):
            Ngay = record.get("Ngay", "")
            GiaDongCua = record.get("GiaDongCua", "")
            GiaDieuChinh = record.get("GiaDieuChinh", "")
            if GiaDieuChinh == GiaDongCua:
                GiaDieuChinh = "--"
            
            ThayDoi = record.get("ThayDoi", "")
            KhoiLuongKhopLenh = record.get("KhoiLuongKhopLenh", 0)
            GiaTriKhopLenh = record.get("GiaTriKhopLenh", 0)
            GiaTriKhopLenh_ti = GiaTriKhopLenh / 1000000000.0
            
            KLThoaThuan = record.get("KLThoaThuan", 0)
            GtThoaThuan = record.get("GtThoaThuan", 0)
            GtThoaThuan_ti = GtThoaThuan / 1000000000.0
            
            GiaMoCua = record.get("GiaMoCua", "")
            GiaCaoNhat = record.get("GiaCaoNhat", "")
            GiaThapNhat = record.get("GiaThapNhat", "")
            
            ws.cell(row=i, column=1).value = Ngay
            ws.cell(row=i, column=2).value = f"{GiaDongCua}"
            ws.cell(row=i, column=3).value = f"{GiaDieuChinh}"
            ws.cell(row=i, column=4).value = f"{ThayDoi}"
            ws.cell(row=i, column=5).value = f"{KhoiLuongKhopLenh:,}"
            ws.cell(row=i, column=6).value = f"{GiaTriKhopLenh_ti:,.2f}"
            ws.cell(row=i, column=7).value = f"{KLThoaThuan:,}"
            ws.cell(row=i, column=8).value = f"{GtThoaThuan_ti:,.2f}"
            ws.cell(row=i, column=9).value = f"{GiaMoCua}"
            ws.cell(row=i, column=10).value = f"{GiaCaoNhat}"
            ws.cell(row=i, column=11).value = f"{GiaThapNhat}"
        
        # Set kích thước cột
        # A,B,C,E,G,I,J,K = 75 pixels ~ width=11
        # D = 95 pixels ~ width=14
        # F,H = 105 pixels ~ width=15
        col_widths = {
            'A': 11, 'B': 11, 'C': 11, 'D': 14, 'E': 11,
            'F': 15, 'G': 11, 'H': 15, 'I': 11, 'J': 11, 'K': 11
        }
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w
        
        # Kẻ viền all border
        last_row = ws.max_row
        last_col = ws.max_column
        
        thin_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        for r in range(1, last_row + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
        
        # Lưu file Excel
        file_name = f"{symbol}_{start_str}-{end_str}.xlsx"
        xlsx_path = os.path.join(xlsx_folder, file_name)
        wb.save(xlsx_path)
        print(f"Đã lưu dữ liệu cho {symbol} vào file {xlsx_path}")

    else:
        print(f"Lỗi khi gọi API cho mã {symbol}, Status code:", response.status_code)

# Sau khi tạo xong tất cả file Excel, gọi hàm convert_all_excel_to_pdf
convert_all_excel_to_pdf(xlsx_folder, pdf_folder)

# Gộp file PDF
output_file_name = f"{start_str}_{end_str}.pdf"
merged_pdf_path = merge_pdfs_in_folder(pdf_folder, root_path, output_file_name)

load_dotenv()

# Lấy các biến môi trường
sender_email = os.getenv("SENDER_EMAIL")
sender_password = os.getenv("SENDER_PASSWORD")
recipient_email = os.getenv("RECIPIENT_EMAIL")

# Gửi email
email_subject = f"Lịch sử giao dịch {start_str} - {end_str}"
email_body = "Xin vui lòng xem các file PDF đính kèm."

send_email_with_attachments(
    pdf_folder=root_path,
    sender_email=sender_email,
    sender_password=sender_password,
    recipient_email=recipient_email,
    subject=email_subject,
    body=email_body
)

pycache = os.path.join(root_path, "__pycache__")
delete_folder(xlsx_folder)
delete_folder(pdf_folder)
delete_folder(pycache)